"""
Live vs Backtest Comparison — Portfolio C.

Reads paper_trade_log.xlsx and compares live/paper results
against frozen backtest reference values.

Reports:
  1. Live R metrics vs backtest reference
  2. Book-level comparison
  3. Robustness checks
  4. Promotion readiness assessment
  5. Equity curve (daily cumulative R)

Usage:
    python -m alert_overlay.live_vs_backtest
"""

import sys
import statistics
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl --break-system-packages")
    sys.exit(1)

TRADE_LOG = Path(__file__).parent / "paper_trade_log.xlsx"

# ── Backtest reference values ──
BT = {
    "n": 79, "wr": 51.9, "pf_r": 1.95, "exp_r": 0.390,
    "total_r": 30.79, "max_dd_r": 6.64, "stop_rate": 17.7,
    "long_n": 45, "long_pf_r": 1.50, "long_exp_r": 0.297,
    "short_n": 34, "short_pf_r": 4.07, "short_exp_r": 0.513,
}

# ── Promotion thresholds ──
PROMO = {
    "min_trades": 50, "min_long": 25, "min_short": 15,
    "min_rt_days": 6, "min_calendar_days": 30, "min_active_days": 15,
    "pf_r_min": 1.3, "exp_r_min": 0.10, "max_dd_max": 15.0,
    "stop_rate_max": 30.0,
    "long_exp_min": 0.0, "short_exp_min": 0.0,
    "long_pf_min": 1.0, "short_pf_min": 1.0,
}


def pf_str(pf):
    return f"{pf:.2f}" if pf < 999 else "inf"


def load_trades():
    """Load trades from paper_trade_log.xlsx Trade Log sheet."""
    if not TRADE_LOG.exists():
        print(f"ERROR: {TRADE_LOG} not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(TRADE_LOG), data_only=True)
    ws = wb["Trade Log"]
    trades = []
    for row in range(2, 202):
        date_val = ws[f"B{row}"].value
        if date_val is None or date_val == "":
            continue
        sym = ws[f"D{row}"].value or ""
        book = ws[f"E{row}"].value or ""
        setup = ws[f"F{row}"].value or ""
        regime = ws[f"G{row}"].value or ""
        pnl_r = ws[f"Q{row}"].value
        exit_reason = ws[f"N{row}"].value or ""
        if pnl_r is None or pnl_r == "":
            continue
        try:
            pnl_r = float(pnl_r)
        except (ValueError, TypeError):
            continue
        trades.append({
            "date": str(date_val)[:10],
            "sym": str(sym),
            "book": str(book).upper(),
            "setup": str(setup),
            "regime": str(regime),
            "pnl_r": pnl_r,
            "exit_reason": str(exit_reason),
        })
    wb.close()
    return trades


def compute_r(trades):
    n = len(trades)
    if n == 0:
        return {"n": 0, "wr": 0, "pf_r": 0, "exp_r": 0, "total_r": 0,
                "max_dd_r": 0, "stop_rate": 0}
    wins = [t for t in trades if t["pnl_r"] > 0]
    losses = [t for t in trades if t["pnl_r"] <= 0]
    total_r = sum(t["pnl_r"] for t in trades)
    gw = sum(t["pnl_r"] for t in wins)
    gl = abs(sum(t["pnl_r"] for t in losses))
    pf_r = gw / gl if gl > 0 else float("inf")
    cum = pk = dd = 0.0
    for t in trades:
        cum += t["pnl_r"]
        if cum > pk:
            pk = cum
        if pk - cum > dd:
            dd = pk - cum
    stopped = sum(1 for t in trades if t["exit_reason"] == "stop")
    return {
        "n": n, "wr": len(wins) / n * 100, "pf_r": pf_r,
        "exp_r": total_r / n, "total_r": total_r,
        "max_dd_r": dd, "stop_rate": stopped / n * 100,
    }


def main():
    trades = load_trades()

    if not trades:
        print("=" * 90)
        print("LIVE vs BACKTEST — NO TRADES LOGGED YET")
        print("=" * 90)
        print("\nStart logging trades in paper_trade_log.xlsx.")
        print("Run this script again after accumulating trades.")
        return

    longs = [t for t in trades if t["book"] == "LONG"]
    shorts = [t for t in trades if t["book"] == "SHORT"]
    m_all = compute_r(trades)
    m_long = compute_r(longs)
    m_short = compute_r(shorts)

    # Unique dates
    all_dates = sorted(set(t["date"] for t in trades))
    rt_days = sorted(set(t["date"] for t in trades if "RED+TREND" in t["regime"]))

    print("=" * 90)
    print("LIVE vs BACKTEST COMPARISON — PORTFOLIO C")
    print("=" * 90)
    print(f"Trades logged: {len(trades)}  |  Long: {len(longs)}  Short: {len(shorts)}")
    print(f"Date range: {all_dates[0]} to {all_dates[-1]}  |  Active days: {len(all_dates)}")
    print(f"RED+TREND days: {len(rt_days)}")

    # ═══════════════════════════════════════════════════════════════
    #  SECTION 1: CORE METRICS COMPARISON
    # ═══════════════════════════════════════════════════════════════
    print("\n" + "=" * 90)
    print("SECTION 1: CORE R METRICS — LIVE vs BACKTEST")
    print("=" * 90)

    def compare_row(label, live_val, bt_val, fmt=".2f", higher_is_better=True):
        delta = live_val - bt_val
        if higher_is_better:
            status = "OK" if live_val >= bt_val * 0.7 else "WARN"
        else:
            status = "OK" if live_val <= bt_val * 1.5 else "WARN"
        print(f"    {label:<20}  Live: {live_val:{fmt}}  BT: {bt_val:{fmt}}  "
              f"Delta: {delta:+{fmt}}  [{status}]")

    print()
    compare_row("N", m_all["n"], BT["n"], "d")
    compare_row("WR", m_all["wr"], BT["wr"], ".1f")
    compare_row("PF(R)", m_all["pf_r"], BT["pf_r"])
    compare_row("Exp(R)", m_all["exp_r"], BT["exp_r"], ".3f")
    compare_row("TotalR", m_all["total_r"], BT["total_r"])
    compare_row("MaxDD(R)", m_all["max_dd_r"], BT["max_dd_r"], ".2f", higher_is_better=False)
    compare_row("StopRate", m_all["stop_rate"], BT["stop_rate"], ".1f", higher_is_better=False)

    # ═══════════════════════════════════════════════════════════════
    #  SECTION 2: BOOK-LEVEL COMPARISON
    # ═══════════════════════════════════════════════════════════════
    print("\n" + "=" * 90)
    print("SECTION 2: BOOK-LEVEL COMPARISON")
    print("=" * 90)

    print(f"\n  Long book:")
    compare_row("N", m_long["n"], BT["long_n"], "d")
    compare_row("PF(R)", m_long["pf_r"], BT["long_pf_r"])
    compare_row("Exp(R)", m_long["exp_r"], BT["long_exp_r"], ".3f")

    print(f"\n  Short book:")
    compare_row("N", m_short["n"], BT["short_n"], "d")
    compare_row("PF(R)", m_short["pf_r"], BT["short_pf_r"])
    compare_row("Exp(R)", m_short["exp_r"], BT["short_exp_r"], ".3f")

    # ═══════════════════════════════════════════════════════════════
    #  SECTION 3: ROBUSTNESS
    # ═══════════════════════════════════════════════════════════════
    print("\n" + "=" * 90)
    print("SECTION 3: ROBUSTNESS CHECKS")
    print("=" * 90)

    # Best day / top symbol
    day_r = defaultdict(float)
    sym_r = defaultdict(float)
    for t in trades:
        day_r[t["date"]] += t["pnl_r"]
        sym_r[t["sym"]] += t["pnl_r"]

    if day_r:
        best_day = max(day_r, key=day_r.get)
        top_sym = max(sym_r, key=sym_r.get)

        no_bd = [t for t in trades if t["date"] != best_day]
        no_ts = [t for t in trades if t["sym"] != top_sym]
        m_nbd = compute_r(no_bd)
        m_nts = compute_r(no_ts)

        print(f"\n    Best day: {best_day} ({day_r[best_day]:+.2f}R)")
        print(f"    Top sym:  {top_sym} ({sym_r[top_sym]:+.2f}R)")
        print(f"    Excl best day: PF(R)={pf_str(m_nbd['pf_r'])}  Exp={m_nbd['exp_r']:+.3f}R  "
              f"[{'PASS' if m_nbd['pf_r'] >= 1.0 else 'FAIL'}]")
        print(f"    Excl top sym:  PF(R)={pf_str(m_nts['pf_r'])}  Exp={m_nts['exp_r']:+.3f}R  "
              f"[{'PASS' if m_nts['pf_r'] >= 1.0 else 'FAIL'}]")

        # Concentration
        profitable_days = sum(1 for v in day_r.values() if v > 0)
        pos_syms = sum(1 for v in sym_r.values() if v > 0)
        total_syms = len(sym_r)
        total_abs_r = m_all["total_r"] if m_all["total_r"] != 0 else 1
        top_sym_pct = sym_r[top_sym] / total_abs_r * 100 if total_abs_r != 0 else 0
        top_day_pct = day_r[best_day] / total_abs_r * 100 if total_abs_r != 0 else 0

        print(f"\n    Profitable days: {profitable_days}/{len(day_r)} "
              f"({profitable_days/len(day_r)*100:.0f}%)")
        print(f"    Positive symbols: {pos_syms}/{total_syms} "
              f"({pos_syms/total_syms*100:.0f}%)")
        print(f"    Top sym concentration: {top_sym_pct:.1f}% of TotalR  "
              f"[{'PASS' if abs(top_sym_pct) < 30 else 'WARN'}]")
        print(f"    Top day concentration: {top_day_pct:.1f}% of TotalR  "
              f"[{'PASS' if abs(top_day_pct) < 30 else 'WARN'}]")

    # ═══════════════════════════════════════════════════════════════
    #  SECTION 4: DAILY EQUITY CURVE
    # ═══════════════════════════════════════════════════════════════
    print("\n" + "=" * 90)
    print("SECTION 4: DAILY EQUITY CURVE (R)")
    print("=" * 90)

    print(f"\n    {'Date':<12}  {'N':>3}  {'Day R':>8}  {'Cum R':>8}")
    print(f"    {'-'*12}  {'-'*3}  {'-'*8}  {'-'*8}")
    cum = 0.0
    for d in sorted(day_r.keys()):
        n_d = sum(1 for t in trades if t["date"] == d)
        cum += day_r[d]
        print(f"    {d}  {n_d:3d}  {day_r[d]:+8.2f}R  {cum:+8.2f}R")

    # ═══════════════════════════════════════════════════════════════
    #  SECTION 5: PROMOTION READINESS
    # ═══════════════════════════════════════════════════════════════
    print("\n" + "=" * 90)
    print("SECTION 5: PROMOTION READINESS")
    print("=" * 90)

    # Sample checks
    sample_checks = [
        ("Total trades ≥ 50", m_all["n"] >= PROMO["min_trades"]),
        ("Long trades ≥ 25", m_long["n"] >= PROMO["min_long"]),
        ("Short trades ≥ 15", m_short["n"] >= PROMO["min_short"]),
        ("RED+TREND days ≥ 6", len(rt_days) >= PROMO["min_rt_days"]),
        ("Calendar days ≥ 30", len(all_dates) >= PROMO["min_calendar_days"]),  # approximate
        ("Active days ≥ 15", len(all_dates) >= PROMO["min_active_days"]),
    ]

    print(f"\n  Sample Requirements:")
    sample_met = all(p for _, p in sample_checks)
    for label, passed in sample_checks:
        print(f"    [{'PASS' if passed else 'WAIT'}] {label}")

    if not sample_met:
        print(f"\n  STATUS: COLLECTING DATA — sample requirements not yet met.")
        print(f"  Continue logging trades. Do not evaluate promotion metrics yet.")
        return

    # Promotion metrics
    print(f"\n  Tier 1: Core Performance:")
    tier1 = [
        ("PF(R) ≥ 1.3", m_all["pf_r"] >= PROMO["pf_r_min"]),
        ("Exp(R) > +0.10R", m_all["exp_r"] > PROMO["exp_r_min"]),
        ("TotalR > 0", m_all["total_r"] > 0),
        ("MaxDD(R) < 15R", m_all["max_dd_r"] < PROMO["max_dd_max"]),
        ("Stop Rate < 30%", m_all["stop_rate"] < PROMO["stop_rate_max"]),
    ]
    for label, passed in tier1:
        print(f"    [{'PASS' if passed else 'FAIL'}] {label}")

    print(f"\n  Tier 2: Book-Level:")
    tier2 = [
        ("Long Exp(R) > 0", m_long["exp_r"] > PROMO["long_exp_min"]),
        ("Short Exp(R) > 0", m_short["exp_r"] > PROMO["short_exp_min"]),
        ("Long PF(R) ≥ 1.0", m_long["pf_r"] >= PROMO["long_pf_min"]),
        ("Short PF(R) ≥ 1.0", m_short["pf_r"] >= PROMO["short_pf_min"]),
    ]
    for label, passed in tier2:
        print(f"    [{'PASS' if passed else 'FAIL'}] {label}")

    print(f"\n  Tier 3: Robustness:")
    tier3 = []
    if day_r:
        tier3.append(("Excl best day PF(R) ≥ 1.0", m_nbd["pf_r"] >= 1.0))
        tier3.append(("Excl top sym PF(R) ≥ 1.0", m_nts["pf_r"] >= 1.0))
        tier3.append(("≥50% days profitable",
                       profitable_days / len(day_r) >= 0.50))
        tier3.append(("≥50% symbols positive R",
                       pos_syms / total_syms >= 0.50 if total_syms > 0 else False))
        tier3.append(("No symbol > 30% of TotalR",
                       abs(top_sym_pct) < 30))
        tier3.append(("No day > 30% of TotalR",
                       abs(top_day_pct) < 30))
    for label, passed in tier3:
        print(f"    [{'PASS' if passed else 'FAIL'}] {label}")

    # Overall
    all_pass = (all(p for _, p in tier1) and all(p for _, p in tier2)
                and all(p for _, p in tier3))
    t1_pass = all(p for _, p in tier1)
    t2_pass = all(p for _, p in tier2)

    print(f"\n  {'='*60}")
    if all_pass:
        print(f"  VERDICT: ALL TIERS PASS — ELIGIBLE FOR PROMOTION")
    elif not t1_pass:
        print(f"  VERDICT: TIER 1 FAIL — System lacks edge. Consider demotion.")
    elif not t2_pass:
        print(f"  VERDICT: TIER 2 FAIL — One book underperforming. Investigate.")
    else:
        print(f"  VERDICT: TIER 3 FAIL — Robustness concerns. Continue collecting data.")
    print(f"  {'='*60}")
    print()


if __name__ == "__main__":
    main()
