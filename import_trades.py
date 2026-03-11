#!/usr/bin/env python3
"""
Import IBKR trades into Trade_Journal.xlsx

Usage:
  # From a local CSV export (like trades.20260306.csv):
  python import_trades.py trades.20260306.csv

  # From IBKR Flex Query (once configured):
  python import_trades.py --flex

  # Preview trades without writing to journal:
  python import_trades.py trades.20260306.csv --dry-run

  # Specify a custom journal path:
  python import_trades.py trades.20260306.csv --journal /path/to/Trade_Journal.xlsx
"""

import argparse
import csv
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, time
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Config ──
SCRIPT_DIR = Path(__file__).parent
DEFAULT_JOURNAL = SCRIPT_DIR / "Trade_Journal.xlsx"
FLEX_CONFIG = SCRIPT_DIR / ".flex_config"

# IBKR Flex Web Service endpoints
FLEX_SEND_URL = "https://ndcdyn.interactivebrokers.com/AccountManagement/FlexWebService/SendRequest"
FLEX_GET_URL = "https://ndcdyn.interactivebrokers.com/AccountManagement/FlexWebService/GetStatement"

# Trade Log column mapping (1-indexed)
COL = {
    "trade_num": 1,   # A
    "date": 2,        # B
    "time_in": 3,     # C
    "symbol": 4,      # D
    "direction": 5,   # E
    "setup": 6,       # F
    "entry": 7,       # G
    "stop": 8,        # H
    "target": 9,      # I
    "shares": 10,     # J
    "exit": 11,       # K
    "exit_time": 12,  # L
    "exit_reason": 13,# M
    "pnl": 14,        # N
    "pnl_pct": 15,    # O
    "r_mult": 16,     # P
    "result": 17,     # Q
    "held_min": 18,   # R
    "notes": 19,      # S
    "screenshot": 20, # T
    "grade": 21,      # U
    "rules": 22,      # V
    "emotion": 23,    # W
    "mistake": 24,    # X
    "regime": 25,     # Y
    "fees": 26,       # Z
    "mae": 27,        # AA
    "mfe": 28,        # AB
    "net_pnl": 29,    # AC
}

# ── Styles ──
BODY_FONT = Font(name="Arial", size=10)
WIN_FONT = Font(name="Arial", size=10, bold=True, color="2E7D32")
LOSS_FONT = Font(name="Arial", size=10, bold=True, color="C62828")
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"))
MONEY_FMT = '$#,##0.00'
PCT_FMT = '0.0%'


def parse_ibkr_csv(filepath):
    """Parse IBKR trade CSV into matched round-trip trades."""
    raw = []
    with open(filepath, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sym = row.get("Symbol", "").strip()
            if not sym:
                continue
            raw.append({
                "symbol": sym,
                "action": row.get("Action", "").strip(),
                "qty": abs(int(float(row.get("Quantity", 0)))),
                "price": float(row.get("Price", 0)),
                "time": row.get("Time", "").strip(),
                "date": row.get("Date", "").strip(),
            })

    # Match BOT/SLD pairs by symbol in order
    opens = {}  # symbol -> list of open legs
    trades = []

    for fill in raw:
        sym = fill["symbol"]
        action = fill["action"]

        if action in ("BOT", "SLD"):
            # Check if this closes an existing position
            opposite = "SLD" if action == "BOT" else "BOT"
            if sym in opens and opens[sym] and opens[sym][0]["action"] == opposite:
                opener = opens[sym].pop(0)
                if not opens[sym]:
                    del opens[sym]

                if opener["action"] == "BOT":
                    direction = "LONG"
                    entry_price = opener["price"]
                    exit_price = fill["price"]
                    entry_time = opener["time"]
                    exit_time = fill["time"]
                else:
                    direction = "SHORT"
                    entry_price = opener["price"]
                    exit_price = fill["price"]
                    entry_time = opener["time"]
                    exit_time = fill["time"]

                qty = opener["qty"]
                if direction == "LONG":
                    pnl = round(qty * (exit_price - entry_price), 2)
                else:
                    pnl = round(qty * (entry_price - exit_price), 2)

                pnl_pct = round((exit_price - entry_price) / entry_price, 6) if direction == "LONG" else round((entry_price - exit_price) / entry_price, 6)

                t_in = datetime.strptime(entry_time, "%H:%M:%S")
                t_out = datetime.strptime(exit_time, "%H:%M:%S")
                hold_min = max(1, int((t_out - t_in).total_seconds() / 60))

                result = "WIN" if pnl > 0 else ("LOSS" if pnl < 0 else "FLAT")

                trades.append({
                    "symbol": sym,
                    "date": opener["date"],
                    "direction": direction,
                    "entry": entry_price,
                    "exit": exit_price,
                    "qty": qty,
                    "pnl": pnl,
                    "pnl_pct": pnl_pct,
                    "time_in": entry_time,
                    "time_out": exit_time,
                    "hold_min": hold_min,
                    "result": result,
                })
            else:
                opens.setdefault(sym, []).append(fill)

    if opens:
        remaining = sum(len(v) for v in opens.values())
        print(f"  Warning: {remaining} unmatched fill(s) — open positions not journaled")

    return trades


def fetch_flex_trades():
    """Fetch trades from IBKR Flex Web Service."""
    try:
        import urllib.request
        import time as _time
    except ImportError:
        print("ERROR: urllib not available")
        sys.exit(1)

    if not FLEX_CONFIG.exists():
        print("ERROR: Flex Query not configured.")
        print(f"Create {FLEX_CONFIG} with two lines:")
        print("  Line 1: Your Flex Token")
        print("  Line 2: Your Query ID")
        print("\nOr run: python import_trades.py --flex-setup")
        sys.exit(1)

    lines = FLEX_CONFIG.read_text().strip().split("\n")
    token = lines[0].strip()
    query_id = lines[1].strip()

    print(f"Requesting Flex Query {query_id}...")
    url = f"{FLEX_SEND_URL}?t={token}&q={query_id}&v=3"
    resp = urllib.request.urlopen(url).read().decode()
    root = ET.fromstring(resp)

    status = root.find(".//Status")
    if status is not None and status.text != "Success":
        err = root.find(".//ErrorMessage")
        print(f"ERROR: {err.text if err is not None else 'Unknown error'}")
        sys.exit(1)

    ref_code = root.find(".//ReferenceCode").text
    print(f"Reference code: {ref_code}, waiting for report...")

    # Poll for result
    for attempt in range(10):
        _time.sleep(3)
        url2 = f"{FLEX_GET_URL}?t={token}&q={ref_code}&v=3"
        resp2 = urllib.request.urlopen(url2).read().decode()

        if resp2.strip().startswith("<"):
            root2 = ET.fromstring(resp2)
            if root2.find(".//Status") is not None:
                continue  # Still processing
            # Got real data — parse XML trades
            return parse_flex_xml(resp2)
        else:
            # CSV response
            tmp = SCRIPT_DIR / ".flex_temp.csv"
            tmp.write_text(resp2)
            trades = parse_ibkr_csv(str(tmp))
            tmp.unlink()
            return trades

    print("ERROR: Timed out waiting for Flex Query response")
    sys.exit(1)


def parse_flex_xml(xml_str):
    """Parse Flex Query XML response into trades."""
    root = ET.fromstring(xml_str)
    trades = []

    for order in root.iter("Trade"):
        sym = order.get("symbol", "")
        action = order.get("buySell", "")
        qty = abs(int(float(order.get("quantity", 0))))
        price = float(order.get("tradePrice", 0))
        dt = order.get("dateTime", "")
        commission = abs(float(order.get("ibCommission", 0)))

        trades.append({
            "symbol": sym,
            "action": "BOT" if action == "BUY" else "SLD",
            "qty": qty,
            "price": price,
            "time": dt.split(";")[1] if ";" in dt else dt.split(",")[1].strip() if "," in dt else "",
            "date": dt.split(";")[0] if ";" in dt else dt.split(",")[0].strip() if "," in dt else "",
            "commission": commission,
        })

    # Re-match into round trips using same logic
    # For now, write temp CSV and reparse
    return trades


def find_next_row(ws):
    """Find the first empty row in Trade Log."""
    for r in range(2, 2000):
        if ws.cell(row=r, column=COL["symbol"]).value is None:
            return r
    return 2


def find_last_trade_num(ws):
    """Find the highest trade number in the journal."""
    max_num = 0
    for r in range(2, 2000):
        val = ws.cell(row=r, column=COL["trade_num"]).value
        if val is None:
            break
        if isinstance(val, (int, float)):
            max_num = max(max_num, int(val))
    return max_num


def format_date(date_str):
    """Normalize date string to YYYY-MM-DD."""
    if not date_str:
        return ""
    for fmt in ["%Y%m%d", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str


def write_trades(trades, journal_path, dry_run=False):
    """Write matched trades into Trade_Journal.xlsx."""
    if not trades:
        print("No trades to import.")
        return

    if dry_run:
        print(f"\n{'#':>3} {'Symbol':<6} {'Dir':<5} {'Qty':>5} {'Entry':>10} {'Exit':>10} {'P&L':>10} {'Result':<5} {'Hold'}")
        print("-" * 70)
        for i, t in enumerate(trades, 1):
            print(f"{i:>3} {t['symbol']:<6} {t['direction']:<5} {t['qty']:>5} {t['entry']:>10.2f} {t['exit']:>10.2f} {t['pnl']:>10.2f} {t['result']:<5} {t['hold_min']}min")
        total = sum(t["pnl"] for t in trades)
        wins = sum(1 for t in trades if t["result"] == "WIN")
        print(f"\n  {len(trades)} trades | {wins}W-{len(trades)-wins}L | Total P&L: ${total:.2f}")
        return

    wb = openpyxl.load_workbook(journal_path)
    tl = wb["Trade Log"]

    start_row = find_next_row(tl)
    trade_num = find_last_trade_num(tl) + 1

    for i, t in enumerate(trades):
        r = start_row + i
        n = trade_num + i

        # Parse time strings to time objects
        try:
            parts = t["time_in"].split(":")
            t_in = time(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            t_in = t["time_in"]

        try:
            parts = t["time_out"].split(":")
            t_out = time(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            t_out = t["time_out"]

        tl.cell(row=r, column=COL["trade_num"], value=n)
        tl.cell(row=r, column=COL["date"], value=format_date(t["date"]))
        tl.cell(row=r, column=COL["time_in"], value=t_in)
        tl.cell(row=r, column=COL["time_in"]).number_format = 'h:mm:ss'
        tl.cell(row=r, column=COL["symbol"], value=t["symbol"])
        tl.cell(row=r, column=COL["direction"], value=t["direction"])
        tl.cell(row=r, column=COL["entry"], value=t["entry"])
        tl.cell(row=r, column=COL["shares"], value=t["qty"])
        tl.cell(row=r, column=COL["exit"], value=t["exit"])
        tl.cell(row=r, column=COL["exit_time"], value=t_out)
        tl.cell(row=r, column=COL["exit_time"]).number_format = 'h:mm:ss'
        tl.cell(row=r, column=COL["pnl"], value=t["pnl"])
        tl.cell(row=r, column=COL["pnl_pct"], value=t["pnl_pct"])
        tl.cell(row=r, column=COL["r_mult"]).value = f'=IF(OR(H{r}="",H{r}=0),"",N{r}/(ABS(G{r}-H{r})*J{r}))'
        tl.cell(row=r, column=COL["result"], value=t["result"])
        tl.cell(row=r, column=COL["held_min"], value=t["hold_min"])
        tl.cell(row=r, column=COL["held_min"]).number_format = '0" min"'
        tl.cell(row=r, column=COL["net_pnl"]).value = f'=IF(N{r}="","",N{r}-IF(Z{r}="",0,Z{r}))'

        # Apply formatting
        for c in range(1, 30):
            cell = tl.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        for col in [COL["entry"], COL["exit"], COL["pnl"], COL["net_pnl"]]:
            tl.cell(row=r, column=col).number_format = MONEY_FMT
        tl.cell(row=r, column=COL["pnl_pct"]).number_format = PCT_FMT

        if t["result"] == "WIN":
            tl.cell(row=r, column=COL["pnl"]).font = WIN_FONT
            tl.cell(row=r, column=COL["result"]).font = WIN_FONT
        elif t["result"] == "LOSS":
            tl.cell(row=r, column=COL["pnl"]).font = LOSS_FONT
            tl.cell(row=r, column=COL["result"]).font = LOSS_FONT

    wb.save(journal_path)

    total = sum(t["pnl"] for t in trades)
    wins = sum(1 for t in trades if t["result"] == "WIN")
    print(f"\nImported {len(trades)} trades into rows {start_row}-{start_row + len(trades) - 1}")
    print(f"  {wins}W-{len(trades)-wins}L | Total P&L: ${total:.2f}")
    print(f"  Saved: {journal_path}")
    print(f"\nRemember to fill in: Setup, Grade, Rules?, Emotion, Mistake, Regime")


def setup_flex():
    """Interactive setup for Flex Query credentials."""
    print("IBKR Flex Query Setup")
    print("=" * 40)
    print("\n1. Log into IBKR Account Management")
    print("2. Go to Reports → Flex Queries")
    print("3. Create a Trade Confirmation Flex Query")
    print("4. Note the Query ID")
    print("5. Go to Reports → Settings → Flex Web Service")
    print("6. Activate and copy your Token\n")
    token = input("Paste your Flex Token: ").strip()
    query_id = input("Paste your Query ID: ").strip()
    FLEX_CONFIG.write_text(f"{token}\n{query_id}\n")
    print(f"\nSaved to {FLEX_CONFIG}")
    print("You can now run: python import_trades.py --flex")


def main():
    parser = argparse.ArgumentParser(description="Import IBKR trades into Trade Journal")
    parser.add_argument("csv_file", nargs="?", help="Path to IBKR trade CSV export")
    parser.add_argument("--flex", action="store_true", help="Fetch trades from IBKR Flex Query")
    parser.add_argument("--flex-setup", action="store_true", help="Configure Flex Query credentials")
    parser.add_argument("--dry-run", action="store_true", help="Preview trades without writing")
    parser.add_argument("--journal", default=str(DEFAULT_JOURNAL), help="Path to Trade_Journal.xlsx")
    args = parser.parse_args()

    if args.flex_setup:
        setup_flex()
        return

    if args.flex:
        trades = fetch_flex_trades()
    elif args.csv_file:
        if not os.path.exists(args.csv_file):
            print(f"ERROR: File not found: {args.csv_file}")
            sys.exit(1)
        print(f"Parsing {args.csv_file}...")
        trades = parse_ibkr_csv(args.csv_file)
    else:
        parser.print_help()
        sys.exit(1)

    print(f"Found {len(trades)} round-trip trades")
    write_trades(trades, args.journal, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
